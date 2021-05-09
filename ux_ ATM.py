from tkinter import *
import tkinter.messagebox as tsmg
from openpyxl import load_workbook
from openpyxl import Workbook
workbook = load_workbook("/home/athiyan/hello_world.xlsx")
workbook.sheetnames
sheet = workbook.active
root=Tk()


cash=(sheet["E2"].value)
pin=(sheet["F2"].value)

def check_pin():    
    a=num.get()
    if(a==""):
        tsmg.showerror("error","Enter your pin")
    elif(atm_pin!=pin):
        tsmg.showerror("Error","invaild pin ")
    else:
        print("wlecome",sheet["B2"].value)
        print("Press 1 to withdraw cash ")
        print("press 2 to check your balance ")
        print("press 3 to to pay in ")
        print("press 4 to exit")
    option=int(input("enter your choice "))
    while(option >= 5):
        print("enter your choice correctly")
        print("Press 1 to withdraw cash ")
        print("press 2 to check your balance ")
        print("press 3 to to pay in ")
        print("press 4 to exit")
        option=int(input("enter your choice "))


root.geometry("700x1000")
root.title("virual ATM")

num=StringVar()

f1=Frame(root)
Label(f1,text="Welcome To VIrual ATM",font="SegoeUI 30 bold",fg="sky blue").pack(padx=5,pady=10)
f1.pack(fill=BOTH)

f2=Frame(root)
Label(f2,text="Enter your PIN",font="SegoeUI 20 bold",fg="teal").pack(padx=5,pady=5)
e1=Entry(f2,textvariable=num,font="SegoeUI 14 bold",fg="black",bg="white",relief=SUNKEN,borderwidth=4,justify="center").pack(ipady=5)
f2.pack(fill=BOTH,padx=5,pady=10)

f3=Frame(root)
Button(f3,text="check pin",command=check_pin,font="SegoeUI 15 bold",fg="purple").pack(padx=20,pady=10,side=LEFT)
f3.pack()



root.mainloop()

from openpyxl import load_workbook
from openpyxl import Workbook
workbook = load_workbook("/home/athiyan/hello_world.xlsx")
workbook.sheetnames
sheet = workbook.active

print("Welcome to virual ATM")
atm_pin=int(input("Enter your pin: "))
cash=(sheet["E2"].value)
pin=(sheet["F2"].value)
 
if(atm_pin!=pin):
    print("you have enter a wrong password")
else:
    print("wlecome",sheet["B2"].value)
    print("Press 1 to withdraw cash ")
    print("press 2 to check your balance ")
    print("press 3 to to pay in ")
    print("press 4 to exit")
    option=int(input("enter your choice "))
    while(option >= 5):
        print("enter your choice correctly")
        print("Press 1 to withdraw cash ")
        print("press 2 to check your balance ")
        print("press 3 to to pay in ")
        print("press 4 to exit")
        option=int(input("enter your choice "))
if(option==1):
    withdraw=int(input("enter the amount you want to withdraw "))
    withdraw_cash=int(cash)-withdraw
    print(f"your reminding cash is {withdraw_cash} ")
    print("thank you")
        
    SystemExit
elif(option==2):
        print("your current balance is",cash)
        print("thank you")
elif(option==3):
        pay_in=int(input("Enter the amount you want to pay in "))
        add_pay_in=pay_in+int(cash)
        print(f"your current balance is {add_pay_in}")
else:
        print("thanks for visitihng or bank")
        exit
workbook.save(filename="hello_world.xlsx")



        

